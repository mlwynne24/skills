from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from shutil import copyfile

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

NOT_IN_TRANSCRIPT = "(not in transcript)"
ERROR_MARKER = "[error]"
DATA_HEADER_ROW = 2
DATA_FIRST_ROW = 3
FILENAME_COL = 2  # column B

_GREY_ITALIC = Font(name="Arial", size=10, italic=True, color="FF888888")
_PLAIN = Font(name="Arial", size=10)
_HEADER_BOLD = Font(name="Arial", size=10, bold=True)
_THEME_BOLD = Font(name="Arial", size=11, bold=True)
_ERROR_FILL = PatternFill(patternType="solid", fgColor="FFFCE4E4")
_THEME_FILL = PatternFill(patternType="solid", fgColor="FFEAECEE")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

_AUTOFIT_CHAR_FACTOR = 1.1
_AUTOFIT_PADDING = 2
_AUTOFIT_MAX = 50  # cap so long answer cells wrap rather than make columns excessively wide
_AUTOFIT_MIN = 8


def _humanize(name: str) -> str:
    """snake_case -> 'Human Readable'. Schema/data keys stay snake_case;
    this is purely for what consultants see in row 2 and the Errors sheet."""
    parts = [p for p in name.split("_") if p]
    return " ".join(p if p.isupper() else p.capitalize() for p in parts)


@dataclass(frozen=True)
class MetadataColumn:
    name: str
    description: str = ""


@dataclass(frozen=True)
class QuestionColumn:
    name: str
    description: str = ""
    theme: str | None = None


@dataclass(frozen=True)
class RowResult:
    filename: str
    data: dict | None = None  # parsed InterviewGrid.model_dump() on success
    error_code: str | None = None
    error_message: str | None = None
    request_id: str | None = None


@dataclass
class SchemaDrift:
    added: list[str]    # in schema, not in xlsx -> auto-migrate
    removed: list[str]  # in xlsx, not in schema -> exit 2
    renamed_hint: list[tuple[str, str]] = ()  # populated if positional match suggests rename


def append_results(
    output_path: Path,
    template_path: Path,
    title: str,
    metadata_cols: list[MetadataColumn],
    question_cols: list[QuestionColumn],
    results: list[RowResult],
) -> SchemaDrift:
    """Idempotent: load existing or seed from template, write headers if new,
    upsert rows by filename, save. Returns schema drift summary; caller decides
    how to react to drift (auto-migrate added cols, abort on removed)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    is_new = not output_path.exists()
    if is_new:
        copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Interviews"

    if is_new:
        _write_headers(ws, title, metadata_cols, question_cols)
        drift = SchemaDrift(added=[], removed=[])
    else:
        drift = _detect_drift(ws, metadata_cols, question_cols)
        if drift.added and not drift.removed:
            _apply_added(ws, drift.added, metadata_cols, question_cols)

    if not drift.removed:
        existing_rows = _index_existing_rows(ws)
        _ensure_errors_sheet(wb)
        for r in results:
            _upsert_row(ws, existing_rows, metadata_cols, question_cols, r)
        _write_errors_sheet(wb, results)
        _apply_confidence_rule(ws, metadata_cols, question_cols)
        _autofit_columns(ws, metadata_cols, question_cols)

    wb.save(output_path)
    return drift


def _write_headers(
    ws: Worksheet,
    title: str,
    metadata_cols: list[MetadataColumn],
    question_cols: list[QuestionColumn],
) -> None:
    ws.cell(row=1, column=FILENAME_COL).value = title
    col = FILENAME_COL + 1  # col C onwards
    if any(q.theme for q in question_cols):
        groups = _theme_groups(question_cols)
        offset = col + len(metadata_cols)
        for theme, count in groups:
            if theme:
                start, end = offset, offset + count * 2 - 1
                ws.merge_cells(
                    start_row=1, start_column=start, end_row=1, end_column=end
                )
                cell = ws.cell(row=1, column=start)
                cell.value = theme
                cell.font = _THEME_BOLD
                cell.fill = _THEME_FILL
                cell.alignment = _WRAP_CENTER
            offset += count * 2

    fn_cell = ws.cell(row=DATA_HEADER_ROW, column=FILENAME_COL)
    fn_cell.value = "Filename"
    fn_cell.font = _HEADER_BOLD
    fn_cell.alignment = _WRAP_TOP
    for i, m in enumerate(metadata_cols):
        c = ws.cell(row=DATA_HEADER_ROW, column=col + i)
        c.value = _humanize(m.name)
        c.font = _HEADER_BOLD
        c.alignment = _WRAP_TOP
    col += len(metadata_cols)
    for q in question_cols:
        c = ws.cell(row=DATA_HEADER_ROW, column=col)
        c.value = _humanize(q.name)
        c.font = _HEADER_BOLD
        c.alignment = _WRAP_TOP
        c2 = ws.cell(row=DATA_HEADER_ROW, column=col + 1)
        c2.value = "Conf."
        c2.font = _HEADER_BOLD
        c2.alignment = _WRAP_TOP
        ws.column_dimensions[get_column_letter(col + 1)].width = 6
        col += 2


def _theme_groups(question_cols: list[QuestionColumn]) -> list[tuple[str | None, int]]:
    groups: list[tuple[str | None, int]] = []
    for q in question_cols:
        if groups and groups[-1][0] == q.theme:
            groups[-1] = (q.theme, groups[-1][1] + 1)
        else:
            groups.append((q.theme, 1))
    return groups


def _detect_drift(
    ws: Worksheet,
    metadata_cols: list[MetadataColumn],
    question_cols: list[QuestionColumn],
) -> SchemaDrift:
    # Headers on disk are humanised; schema names stay snake_case.
    # Drift is reported in schema-name terms so callers can act on them.
    existing_display = set(_read_header_names(ws))
    schema_names = [m.name for m in metadata_cols] + [q.name for q in question_cols]
    display_to_schema = {_humanize(n): n for n in schema_names}
    added = [n for n in schema_names if _humanize(n) not in existing_display]
    removed = [
        display_to_schema.get(d, d)
        for d in existing_display
        if d not in display_to_schema
    ]
    return SchemaDrift(added=added, removed=removed)


def _read_header_names(ws: Worksheet) -> list[str]:
    names: list[str] = []
    col = FILENAME_COL + 1
    while True:
        v = ws.cell(row=DATA_HEADER_ROW, column=col).value
        if v is None:
            break
        if v != "Conf.":
            names.append(v)
        col += 1
    return names


def _apply_added(
    ws: Worksheet,
    added: list[str],
    metadata_cols: list[MetadataColumn],
    question_cols: list[QuestionColumn],
) -> None:
    # Append added columns at the right edge so existing data stays put.
    # Themes are not retro-merged here; consultants can fix banding manually.
    last_col = ws.max_column
    by_name = {q.name: q for q in question_cols}
    meta_names = {m.name for m in metadata_cols}
    for name in added:
        last_col += 1
        c = ws.cell(row=DATA_HEADER_ROW, column=last_col)
        c.value = _humanize(name)
        c.font = _HEADER_BOLD
        c.alignment = _WRAP_TOP
        if name in by_name:
            last_col += 1
            c2 = ws.cell(row=DATA_HEADER_ROW, column=last_col)
            c2.value = "Conf."
            c2.font = _HEADER_BOLD
            c2.alignment = _WRAP_TOP
        elif name in meta_names:
            pass


def _index_existing_rows(ws: Worksheet) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in range(DATA_FIRST_ROW, ws.max_row + 1):
        v = ws.cell(row=r, column=FILENAME_COL).value
        if v:
            out[str(v)] = r
    return out


def read_processed_filenames(output_path: Path) -> tuple[set[str], set[str]]:
    """Return (clean_set, error_set) of filenames present in the existing grid.
    Empty sets if the file does not exist. A row is classed as an error row when
    its first data cell (col C) holds the [error] marker."""
    if not output_path.exists():
        return set(), set()
    wb = load_workbook(output_path, read_only=True, data_only=True)
    try:
        ws = wb["Interviews"] if "Interviews" in wb.sheetnames else wb.active
        clean: set[str] = set()
        errored: set[str] = set()
        first_data_col = FILENAME_COL + 1
        for r in range(DATA_FIRST_ROW, ws.max_row + 1):
            fn = ws.cell(row=r, column=FILENAME_COL).value
            if not fn:
                continue
            first_val = ws.cell(row=r, column=first_data_col).value
            if first_val == ERROR_MARKER:
                errored.add(str(fn))
            else:
                clean.add(str(fn))
    finally:
        wb.close()
    return clean, errored


@dataclass(frozen=True)
class FilterSummary:
    new: list[str]
    retry: list[str]
    skipped_done: list[str]
    skipped_errors: list[str]


def filter_against_grid(
    files: list[Path],
    output_path: Path,
    retry_errors: bool,
) -> tuple[list[Path], FilterSummary]:
    """Drop inputs already represented in the grid. Files in clean rows are always
    skipped; files in error rows are retried iff retry_errors is true. Returns the
    remaining work list plus a summary so the caller can report what was skipped."""
    clean, errored = read_processed_filenames(output_path)
    new: list[Path] = []
    retry: list[Path] = []
    skipped_done: list[str] = []
    skipped_errors: list[str] = []
    for p in files:
        if p.name in clean:
            skipped_done.append(p.name)
        elif p.name in errored:
            if retry_errors:
                retry.append(p)
            else:
                skipped_errors.append(p.name)
        else:
            new.append(p)
    summary = FilterSummary(
        new=[p.name for p in new],
        retry=[p.name for p in retry],
        skipped_done=skipped_done,
        skipped_errors=skipped_errors,
    )
    return new + retry, summary


def _upsert_row(
    ws: Worksheet,
    existing_rows: dict[str, int],
    metadata_cols: list[MetadataColumn],
    question_cols: list[QuestionColumn],
    r: RowResult,
) -> None:
    row = existing_rows.get(r.filename)
    if row is None:
        row = max([DATA_FIRST_ROW - 1, *existing_rows.values()], default=DATA_FIRST_ROW - 1) + 1
        existing_rows[r.filename] = row
    ws.cell(row=row, column=FILENAME_COL).value = r.filename
    ws.cell(row=row, column=FILENAME_COL).font = _PLAIN
    col = FILENAME_COL + 1

    if r.error_code:
        for _ in metadata_cols:
            cell = ws.cell(row=row, column=col)
            cell.value = ERROR_MARKER
            cell.font = _PLAIN
            cell.fill = _ERROR_FILL
            cell.alignment = _WRAP_TOP
            col += 1
        for _ in question_cols:
            for _shift in (0, 1):
                cell = ws.cell(row=row, column=col + _shift)
                cell.value = ERROR_MARKER if _shift == 0 else None
                cell.font = _PLAIN
                cell.fill = _ERROR_FILL
                cell.alignment = _WRAP_TOP
            col += 2
        return

    data = r.data or {}
    for m in metadata_cols:
        cell = ws.cell(row=row, column=col)
        v = data.get(m.name)
        cell.value = v if v else None
        cell.font = _PLAIN
        cell.fill = PatternFill(fill_type=None)
        cell.alignment = _WRAP_TOP
        col += 1
    for q in question_cols:
        ans = (data.get(q.name) or {}) if isinstance(data.get(q.name), dict) else {}
        value = ans.get("value")
        conf = ans.get("confidence")
        ans_cell = ws.cell(row=row, column=col)
        conf_cell = ws.cell(row=row, column=col + 1)
        ans_cell.fill = PatternFill(fill_type=None)
        conf_cell.fill = PatternFill(fill_type=None)
        ans_cell.alignment = _WRAP_TOP
        conf_cell.alignment = _WRAP_TOP
        if value is None:
            ans_cell.value = NOT_IN_TRANSCRIPT
            ans_cell.font = _GREY_ITALIC
            conf_cell.value = None
            conf_cell.font = _PLAIN
        else:
            ans_cell.value = value
            ans_cell.font = _PLAIN
            if conf is not None:
                conf_cell.value = float(conf)
                conf_cell.number_format = "0%"
            else:
                conf_cell.value = None
            conf_cell.font = _PLAIN
        col += 2


def _autofit_columns(
    ws: Worksheet,
    metadata_cols: list[MetadataColumn],
    question_cols: list[QuestionColumn],
) -> None:
    """Approximate Excel's double-click-auto-fit. Confidence cols stay narrow
    (already set in _write_headers); col A is the brand spacer; row 1 (brand
    banner / merged theme cells) is skipped so merged values don't drive width."""
    confidence_idxs = set()
    first_q = FILENAME_COL + 1 + len(metadata_cols)
    for i in range(len(question_cols)):
        confidence_idxs.add(first_q + 1 + i * 2)

    last_col = FILENAME_COL + len(metadata_cols) + len(question_cols) * 2
    for col_idx in range(FILENAME_COL, last_col + 1):
        if col_idx in confidence_idxs:
            continue
        max_len = 0
        for row_idx in range(DATA_HEADER_ROW, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            longest = max((len(line) for line in str(v).splitlines()), default=0)
            if longest > max_len:
                max_len = longest
        if max_len == 0:
            continue
        width = min(
            max(max_len * _AUTOFIT_CHAR_FACTOR + _AUTOFIT_PADDING, _AUTOFIT_MIN),
            _AUTOFIT_MAX,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_confidence_rule(
    ws: Worksheet,
    metadata_cols: list[MetadataColumn],
    question_cols: list[QuestionColumn],
) -> None:
    if not question_cols:
        return
    first_q = FILENAME_COL + 1 + len(metadata_cols)
    rule = ColorScaleRule(
        start_type="num", start_value=0, start_color="FFF8696B",
        mid_type="num", mid_value=0.5, mid_color="FFFFEB84",
        end_type="num", end_value=1, end_color="FF63BE7B",
    )
    for i in range(len(question_cols)):
        col = first_q + 1 + i * 2  # confidence col is +1 within each pair
        letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{letter}{DATA_FIRST_ROW}:{letter}1048576", rule
        )


def _ensure_errors_sheet(wb) -> None:
    if "Errors" not in wb.sheetnames:
        s = wb.create_sheet("Errors")
        for i, h in enumerate(["Filename", "Error code", "Error message", "Request ID", "Timestamp"]):
            c = s.cell(row=1, column=i + 1)
            c.value = h
            c.font = _HEADER_BOLD


def _write_errors_sheet(wb, results: list[RowResult]) -> None:
    s = wb["Errors"]
    s.delete_rows(2, s.max_row)
    now = datetime.now().isoformat(timespec="seconds")
    row = 2
    for r in results:
        if not r.error_code:
            continue
        s.cell(row=row, column=1).value = r.filename
        s.cell(row=row, column=2).value = r.error_code
        s.cell(row=row, column=3).value = r.error_message
        s.cell(row=row, column=4).value = r.request_id
        s.cell(row=row, column=5).value = now
        row += 1
